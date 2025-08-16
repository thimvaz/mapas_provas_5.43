# mapa_salas_streamlit.py

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
import random
import io
from datetime import date

# ------------------------------
# Funções auxiliares
# ------------------------------

def carregar_dados_excel(arquivo):
    """
    Carrega dados de alunos e salas do arquivo Excel.

    Args:
        arquivo: O arquivo Excel enviado pelo usuário.

    Returns:
        Uma tupla contendo os DataFrames dos alunos, o Workbook e uma lista de nomes de salas.
    """
    try:
        # Carrega as abas de alunos usando os nomes corretos
        df_alunos1 = pd.read_excel(arquivo, sheet_name="alunos_1")
        df_alunos2 = pd.read_excel(arquivo, sheet_name="alunos_2")

        # Carrega o workbook completo para obter os nomes das abas de sala
        wb = load_workbook(arquivo)
        abas_salas = wb.sheetnames[:-2]
        
        return df_alunos1, df_alunos2, wb, abas_salas
    except Exception as e:
        st.error(f"Erro ao carregar o arquivo Excel. Verifique se as abas 'alunos_1' e 'alunos_2' existem. Erro: {e}")
        return None, None, None, None

def filtrar_e_preparar_alunos(df):
    """
    Filtra alunos com Flex != 1, seleciona as colunas relevantes e os converte para uma lista de dicionários.
    
    Args:
        df (pd.DataFrame): O DataFrame de alunos.

    Returns:
        Uma lista de dicionários, cada um representando um aluno com as colunas necessárias.
    """
    # Filtra os alunos que não são 'Flex'
    df_filtrado = df[df["Flex"] != 1].copy()
    # Embaralha a ordem dos alunos
    df_filtrado = df_filtrado.sample(frac=1).reset_index(drop=True)
    # Seleciona as colunas desejadas para o mapa e a lista final
    colunas_desejadas = ["nome", "turma", "RM", "numero"]
    df_final = df_filtrado[colunas_desejadas]
    # Converte o DataFrame para uma lista de dicionários para manipulação mais eficiente
    return df_final.to_dict('records')

def gerar_mapas_todas_salas(alunos1, alunos2, wb, abas_salas, posicoes_retiradas_por_sala):
    """
    Gera o mapa de alocação para todas as salas.

    Args:
        alunos1 (list): Lista de alunos do grupo 1.
        alunos2 (list): Lista de alunos do grupo 2.
        wb: O Workbook do Excel.
        abas_salas (list): Lista de nomes das abas de sala.
        posicoes_retiradas_por_sala (dict): Dicionário com as posições a serem retiradas.

    Returns:
        Um dicionário onde a chave é o nome da sala e o valor é o mapa de alocação.
    """
    mapas = {}
    
    # Faz cópias das listas de alunos para que a alocação possa ser gerada novamente
    # sem re-carregar ou re-embaralhar os dados originais.
    alunos_g1_para_alocar = list(alunos1)
    alunos_g2_para_alocar = list(alunos2)

    for sala_nome in abas_salas:
        aba_sala = wb[sala_nome]
        n_linhas = int(aba_sala["A2"].value)
        n_colunas = int(aba_sala["B2"].value)
        posicoes_retiradas = posicoes_retiradas_por_sala.get(sala_nome, [])

        mapa = [[None for _ in range(n_colunas)] for _ in range(n_linhas)]
        
        # Itera sobre as colunas da sala para alternar os grupos de alunos
        for col in range(n_colunas):
            lista_alunos = alunos_g1_para_alocar if col % 2 == 0 else alunos_g2_para_alocar
            
            # Itera sobre as linhas dentro da coluna
            for lin in range(n_linhas):
                posicao_atual = (lin, col)
                
                if posicao_atual not in posicoes_retiradas and lista_alunos:
                    aluno = lista_alunos.pop(0)
                    mapa[lin][col] = {
                        "nome": aluno.get("nome", "Nome não disponível"),
                        "turma": aluno.get("turma", "Turma não disponível"),
                        "RM": aluno.get("RM", "RM não disponível"),
                        "numero": aluno.get("numero", "Número não disponível")
                    }

        mapas[sala_nome] = mapa
    
    if alunos_g1_para_alocar or alunos_g2_para_alocar:
        st.warning(f"Atenção: Sobraram {len(alunos_g1_para_alocar)} alunos do Grupo 1 e "
                   f"{len(alunos_g2_para_alocar)} alunos do Grupo 2 que não foram alocados.")

    return mapas

def exibir_mapa_sala(mapa, sala_nome, data_avaliacao):
    """
    Exibe o mapa de uma sala com formatação para cópia e impressão.
    
    Args:
        mapa (list): O mapa de alocação da sala.
        sala_nome (str): O nome da sala.
        data_avaliacao (date): A data da avaliação.
    """
    # Formata a data para exibição
    data_formatada = data_avaliacao.strftime('%d/%m/%Y')
    
    st.write(f"### {sala_nome} - {data_formatada}")

    css = """
    <style>
    .grid-mapa {
        border-collapse: collapse;
        margin: 20px auto;
        font-family: Arial, sans-serif;
        max-width: 1260px;
    }
    .grid-mapa td {
        width: 120px;
        height: 100px;
        border: 1px solid #333;
        text-align: center;
        vertical-align: middle;
        padding: 5px;
    }
    .assento {
        display: flex;
        flex-direction: column;
        justify-content: center;
        align-items: center;
        height: 100%;
        font-size: 12px;
        overflow: hidden;
    }
    .assento small {
        font-size: 9px;
        color: #666;
    }
    </style>
    """

    html = "<div class='print-area'><table class='grid-mapa'>"
    for linha in mapa:
        html += "<tr>"
        for assento in linha:
            if assento is None:
                html += "<td></td>"
            else:
                nome = assento.get("nome", "N/A")
                turma = assento.get("turma", "N/A")
                html += f"<td><div class='assento'><strong>{nome}</strong><br><small>{turma}</small></div></td>"
        html += "</tr>"
    html += "</table></div>"

    st.markdown(css + html, unsafe_allow_html=True)

def gerar_lista_por_turma_global(mapas, data_avaliacao):
    """
    Gera uma lista global de todos os alunos alocados com suas respectivas salas e posições.
    
    Args:
        mapas (dict): Dicionário de mapas de sala.
        data_avaliacao (date): A data da avaliação.

    Returns:
        Um DataFrame do Pandas com a lista de alunos e suas alocações.
    """
    dados = []
    for sala_nome, mapa in mapas.items():
        for lin, linha in enumerate(mapa):
            for col, assento in enumerate(linha):
                if assento is not None:
                    dados.append({
                        "turma": assento.get("turma", "N/A"),
                        "nome": assento.get("nome", "N/A"),
                        "RM": assento.get("RM", "N/A"),
                        "numero": assento.get("numero", "N/A"),
                        "sala": sala_nome,
                        "linha": lin + 1,
                        "coluna": col + 1,
                        "data_avaliacao": data_avaliacao.strftime('%d/%m/%Y')
                    })
    return pd.DataFrame(dados)

def to_excel_bytes(df):
    """
    Converte um DataFrame para um objeto BytesIO no formato Excel.
    
    Args:
        df (pd.DataFrame): O DataFrame a ser convertido.

    Returns:
        Um objeto BytesIO com os dados do DataFrame em formato XLSX.
    """
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Alunos Alocados')
    output.seek(0)
    return output

# ------------------------------
# Streamlit Interface
# ------------------------------
st.set_page_config(page_title="Mapas de Sala", layout="wide")
st.title("📝 Assistente Para Mapas de Provas")

st.sidebar.markdown("# Assistente de Provas")
# Substitui a imagem placeholder por um arquivo local e utiliza o parâmetro correto para ajuste de largura
st.sidebar.image("logo.png", use_container_width=True)
st.sidebar.header("Configuração")

arquivo = st.sidebar.file_uploader("Envie o arquivo Excel", type=[".xlsx"])
data_avaliacao = st.sidebar.date_input("Data da Avaliação", date.today())

if arquivo:
    df_1_raw, df_2_raw, wb, abas_salas = carregar_dados_excel(arquivo)
    
    if wb and abas_salas:
        alunos1_lista = filtrar_e_preparar_alunos(df_1_raw)
        alunos2_lista = filtrar_e_preparar_alunos(df_2_raw)

        # Exibe o total de alunos
        st.info(f"Total de alunos para alocar: {len(alunos1_lista)} (Grupo 1) e {len(alunos2_lista)} (Grupo 2).")
        
        # Exibe os detalhes das salas antes da configuração de carteiras
        st.subheader("Capacidade das Salas")
        for sala_nome in abas_salas:
            aba_sala = wb[sala_nome]
            n_linhas = int(aba_sala["A2"].value)
            n_colunas = int(aba_sala["B2"].value)
            total_lugares = n_linhas * n_colunas
            st.markdown(f"**{sala_nome}:** {total_lugares} lugares | {n_linhas} linhas x {n_colunas} colunas")
        
        st.markdown("---")

        posicoes_retiradas_por_sala = {}
        for sala_nome in abas_salas:
            st.subheader(f"Configuração de Carteiras - {sala_nome}")
            aba_sala = wb[sala_nome]
            n_linhas = int(aba_sala["A2"].value)
            n_colunas = int(aba_sala["B2"].value)
            
            posicoes_retiradas = []
            cols_checkbox = st.columns(n_colunas)
            
            for lin in range(n_linhas):
                for col in range(n_colunas):
                    marcado = cols_checkbox[col].checkbox("", key=f"{sala_nome}_ret_{lin}_{col}")
                    if marcado:
                        posicoes_retiradas.append((lin, col))
            
            posicoes_retiradas_por_sala[sala_nome] = posicoes_retiradas

        if st.button("🎓 Gerar Mapas de Todas as Salas"):
            mapas = gerar_mapas_todas_salas(alunos1_lista, alunos2_lista, wb, abas_salas, posicoes_retiradas_por_sala)

            st.header("Mapas de Alocação Gerados")
            for sala_nome, mapa in mapas.items():
                exibir_mapa_sala(mapa, sala_nome, data_avaliacao)
                st.write("---")

            # Geração da lista global
            st.subheader("📄 Lista de Alocação Global")
            lista_global = gerar_lista_por_turma_global(mapas, data_avaliacao)
            st.dataframe(lista_global, use_container_width=True)
            st.download_button(
                "🔗 Baixar Lista Global (XLSX)",
                data=to_excel_bytes(lista_global),
                file_name="lista_de_alocacao_global.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

            # Geração das listas por turma
            st.write("---")
            st.subheader("📄 Listas de Alocação por Turma")
            todas_turmas = sorted(lista_global["turma"].unique())
            for turma_nome in todas_turmas:
                with st.expander(f"Alunos da Turma: {turma_nome}"):
                    df_turma = lista_global[lista_global["turma"] == turma_nome].reset_index(drop=True)
                    st.dataframe(df_turma, use_container_width=True)
                    st.download_button(
                        f"🔗 Baixar Lista da Turma {turma_nome} (XLSX)",
                        data=to_excel_bytes(df_turma),
                        file_name=f"lista_de_alocacao_{turma_nome}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key=f"download_{turma_nome}"
                    )

            # Geração das listas por sala
            st.write("---")
            st.subheader("📄 Listas de Alocação por Sala")
            todas_salas = sorted(lista_global["sala"].unique())
            for sala_nome in todas_salas:
                with st.expander(f"Alunos da Sala: {sala_nome}"):
                    df_sala = lista_global[lista_global["sala"] == sala_nome].reset_index(drop=True)
                    st.dataframe(df_sala, use_container_width=True)
                    st.download_button(
                        f"🔗 Baixar Lista da Sala {sala_nome} (XLSX)",
                        data=to_excel_bytes(df_sala),
                        file_name=f"lista_de_alocacao_sala_{sala_nome}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key=f"download_sala_{sala_nome}"
                    )

